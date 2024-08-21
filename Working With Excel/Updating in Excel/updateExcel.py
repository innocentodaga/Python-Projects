# The program does the following:
# •	 Loops over all the rows.
# •	 If the row is for garlic, celery, or lemons, changes the price.
# 280 
# This means the code will need to do the following:
# •	 Open the spreadsheet file.
# •	 For each row, check whether the value in column A is Celery, Garlic, 
# or Lemon.
# •	 If it is, update the price in column B.
# •	 Save the spreadsheet to a new file (so that you don’t lose the old spreadsheet, just in case).

import openpyxl

# Load the file we want to work on
workbook = openpyxl.load_workbook("./produceSales.xlsx")

# specify the sheet we want to work on
sheet = workbook["Sheet"]

# The produce type  s and their updated prices
# You will only need to update the prices in the lines below 
price_updates = {
    "Garlic" : 3.07,
    "Celery" : 1.19,
    "Lemon" : 1.27,
}

# Loop through the rows and upate the prices
for rowNum in range(2, sheet.max_row): # skip the first row
    # returns the values in the specified cells starting from the first cell to the last one 
    produceName = sheet.cell(row=rowNum, column=1).value   

    # this checks if the values(produceName) is in the dict (price_updates) then the price will be corrected
    if produceName in price_updates:
        sheet.cell(row=rowNum, column=2).value = price_updates[produceName]

# save the updated file as a seperate file
workbook.save("updatedproducesales.xlsx")








       